"""
Author: Martin Bacigal, 01/2025 @ https://procureai.tech
License: MIT License
"""
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from threading import Thread
from dataclasses import dataclass
from typing import Optional, Dict, List, Any, Union
from pathlib import Path
import os
import asyncio
from concurrent.futures import ThreadPoolExecutor
import logging
from functools import partial, lru_cache
import json
from enum import Enum
import mimetypes
import queue
from datetime import datetime
import psutil

import pandas as pd
from pypdf import PdfReader
import docx
import mammoth
from pyxlsb import open_workbook as open_xlsb
import pythoncom
from win32com.client import Dispatch
from openai import AzureOpenAI
from tenacity import retry, stop_after_attempt, wait_exponential
from dataclasses import dataclass, field
from functools import lru_cache
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Enhanced Logging Setup
class LogHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.queue = queue.Queue()
        self.text_widget.tag_configure('ERROR', foreground='red')
        self.text_widget.tag_configure('WARNING', foreground='orange')
        self.text_widget.tag_configure('INFO', foreground='green')
        self.update_widget()

    def emit(self, record):
        self.queue.put(record)

    def update_widget(self):
        while True:
            try:
                record = self.queue.get_nowait()
                self.text_widget.insert('end',
                    f"{datetime.now().strftime('%H:%M:%S')} - {record.getMessage()}\n",
                    record.levelname)
                self.text_widget.see('end')
            except queue.Empty:
                break
        self.text_widget.after(100, self.update_widget)

# Configuration Classes
@dataclass(frozen=True)
class AzureConfig:
    endpoint: str = field(default_factory=lambda: os.getenv("AZURE_ENDPOINT", "AZURE_ENDPOINT_URL"))
    deployment: str = field(default_factory=lambda: os.getenv("AZURE_DEPLOYMENT", "AZURE_DEVELOPMENT"))
    api_key: str = field(default_factory=lambda: os.getenv("AZURE_API_KEY", "AZURE_API_KEY"))
    api_version: str = "2024-05-01-preview"

    @lru_cache(maxsize=1)
    def create_client(self):
        if not self.endpoint or not self.api_key:
            logging.error("Azure endpoint or API key is not set.")
            raise ValueError("Azure endpoint or API key is not set.")
       
        try:
            client = AzureOpenAI(
                azure_endpoint=self.endpoint,
                api_key=self.api_key,
                api_version=self.api_version
            )
            logging.info("AzureOpenAI client created successfully.")
            return client
        except Exception as e:
            logging.error(f"Failed to create AzureOpenAI client: {e}")
            raise

@dataclass
class ProcessingConfig:
    MAX_WORKERS: int = os.cpu_count() or 4
    BATCH_SIZE: int = 4
    OUTPUT_DIR: Path = Path("output")
    LOG_LEVEL: str = "DEBUG"

class DocumentType(Enum):
    PDF = "application/pdf"
    DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    DOC = "application/msword"
    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    XLS = "application/vnd.ms-excel"

class ContentAnalyzer:
    def __init__(self, azure_config: AzureConfig):
        self.client = azure_config.create_client()
        self.deployment = azure_config.deployment
        self.analysis_templates = {
            "contract_details": {
                "system": """Analyze the contract and extract the following information:
                1. Start date: in the format DD/MM/YYYY
                2. End date: in the format DD/MM/YYYY
                3. Contract duration
                4. Key deliverables: per each key deliverable start in new line and with "- "
               
                Make sure that you always reply for each point within the same line.
                Avoid markdown, as the result is used for further processing.
                """,
                "max_tokens": 300
            },
            "vendor_assessment": {
                "system": """Evaluate the vendor document and provide:
                1. Vendor capabilities
                2. Risk factors
                3. Compliance status
                4. Past performance metrics
                5. Key strengths and weaknesses""",
                "max_tokens": 400
            },
            "technical_specs": {
                "system": """Extract and analyze technical specifications including:
                1. Technical requirements
                2. Performance metrics
                3. Compatibility requirements
                4. Implementation timeline
                5. Maintenance requirements""",
                "max_tokens": 350
            }
        }

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
    async def analyze_content(self, text: str, analysis_type: str) -> Dict[str, Any]:
        template = self.analysis_templates.get(analysis_type)
        if not template:
            raise ValueError(f"Unknown analysis type: {analysis_type}")

        try:
            completion = await asyncio.to_thread(
                self.client.chat.completions.create,
                model=self.deployment,
                messages=[
                    {"role": "system", "content": template["system"]},
                    {"role": "user", "content": text}
                ],
                max_tokens=template["max_tokens"],
                temperature=0.7
            )
        except Exception as e:
            logging.error(f"Error analyzing content: {e}")
            raise

        return {
            "analysis_type": analysis_type,
            "result": completion.choices[0].message.content
        }

    def format_information(self, result_content: str, analysis_type: str) -> Dict[str, Any]:
        formatted_info = {}
        key_deliverables = []
        lines = result_content.split('\n')
        key_deliverables_flag = False

        for line in lines:
            if "Start date" in line:
                formatted_info["start_date"] = line.split(":")[1].strip()
            elif "End date" in line:
                formatted_info["end_date"] = line.split(":")[1].strip()
            elif "Contract duration" in line:
                formatted_info["contract_duration"] = line.split(":")[1].strip()
            elif "Key deliverables" in line:
                key_deliverables_flag = True
            elif "Vendor capabilities" in line:
                formatted_info["vendor_capabilities"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Risk factors" in line:
                formatted_info["risk_factors"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Compliance status" in line:
                formatted_info["compliance_status"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Past performance metrics" in line:
                formatted_info["past_performance_metrics"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Key strengths and weaknesses" in line:
                formatted_info["key_strengths_weaknesses"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Technical requirements" in line:
                formatted_info["technical_requirements"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Performance metrics" in line:
                formatted_info["performance_metrics"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Compatibility requirements" in line:
                formatted_info["compatibility_requirements"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Implementation timeline" in line:
                formatted_info["implementation_timeline"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif "Maintenance requirements" in line:
                formatted_info["maintenance_requirements"] = line.split(":")[1].strip()
                key_deliverables_flag = False
            elif key_deliverables_flag:
                if line.strip().startswith("-"):
                    key_deliverables.append(line.strip().lstrip("-").strip())

        if key_deliverables:
            formatted_info["key_deliverables"] = key_deliverables

        return formatted_info

class DocumentProcessor:
    def __init__(self, config: ProcessingConfig, azure_config: AzureConfig):
        self.config = config
        self.azure_config = azure_config
        self.analyzer = ContentAnalyzer(azure_config)
        self.executor = ThreadPoolExecutor(max_workers=config.MAX_WORKERS)
       
        # Define processors for different MIME types
        self.processors = {
            DocumentType.PDF.value: self._extract_pdf_content,
            DocumentType.DOCX.value: self._extract_word_content,
            DocumentType.DOC.value: self._extract_word_content,
            DocumentType.XLSX.value: self._extract_excel_content,
            DocumentType.XLS.value: self._extract_excel_content
        }

    async def process_document(self, file_path: Path) -> Dict[str, Any]:
        mime_type, _ = mimetypes.guess_type(str(file_path))
        processor = self.processors.get(mime_type)

        if not processor:
            logging.error(f"Unsupported file type: {mime_type}")
            return {"error": f"Unsupported file type: {mime_type}"}

        try:
            content = await processor(file_path, mime_type)
            if not content:
                logging.error(f"Error extracting content from {file_path}")
                return {"error": f"Error extracting content from {file_path}"}

            analyses = await asyncio.gather(
                self.analyzer.analyze_content(content, "contract_details"),
                self.analyzer.analyze_content(content, "vendor_assessment"),
                self.analyzer.analyze_content(content, "technical_specs")
            )

            return {
                "file_path": str(file_path),
                "metadata": {"mime_type": mime_type},
                "analyses": analyses
            }
        except Exception as e:
            logging.error(f"Processing error for {file_path}: {e}")
            return {"error": f"Processing error for {file_path}: {e}"}

    async def _extract_pdf_content(self, file_path: Path, mime_type: str) -> Optional[str]:
        try:
            reader = PdfReader(str(file_path))
            text = ""
            for page in reader.pages:
                text += page.extract_text()
            return text
        except Exception as e:
            logging.error(f"Error extracting PDF content from {file_path}: {e}")
            return None

    async def _extract_word_content(self, file_path: Path, mime_type: str) -> Optional[str]:
        try:
            if mime_type == DocumentType.DOCX.value:
                doc = docx.Document(str(file_path))
                return "\n".join([para.text for para in doc.paragraphs])
            elif mime_type == DocumentType.DOC.value:
                with open(file_path, "rb") as doc_file:
                    result = mammoth.extract_raw_text(doc_file)
                    return result.value
            return None
        except Exception as e:
            logging.error(f"Error extracting Word content from {file_path}: {e}")
            return None

    async def _extract_excel_content(self, file_path: Path) -> Optional[str]:
        try:
            # Implement extraction logic for Excel files (XLSX, XLS)
            # This is a placeholder implementation and should be replaced with actual logic
            return "Extracted Excel content"
        except Exception as e:
            logging.error(f"Error extracting Excel content from {file_path}: {e}")
            return None

    async def process_directory(self, directory: Path, update_progress, selected_file_types: List[str], update_status, batch_size: int = 10) -> List[Dict[str, Any]]:
        files = [f for f in directory.rglob("*") if f.is_file() and
                 mimetypes.guess_type(str(f))[0] in selected_file_types]

        results = []
        total_files = len(files)
        processed = 0

        output_path = self.config.OUTPUT_DIR / "processed_results.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        for batch in [files[i:i + batch_size] for i in range(0, len(files), batch_size)]:
            batch_results = await asyncio.gather(
                *[self.process_document(f) for f in batch]
            )
            batch_results = [r for r in batch_results if r]  # Filter out None results

            await ResultsExporter.to_excel(batch_results, output_path, self.analyzer, update_progress)

            results.extend(batch_results)
            processed += len(batch)
            update_progress(processed, total_files)

        return results

class EnhancedDocumentProcessorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Enhanced Document Processor")
        self.geometry("800x600")
        self.configure(bg='#f0f0f0')

        # Configurations
        self.azure_config = AzureConfig()
        self.processing_config = ProcessingConfig()

        # File type selection variables
        self.file_types = {
            "PDF": tk.BooleanVar(value=True),
            "DOCX": tk.BooleanVar(value=True),
            "DOC": tk.BooleanVar(value=True),
            "XLSX": tk.BooleanVar(value=True),
            "XLS": tk.BooleanVar(value=True)
        }

        self._setup_styles()
        self._create_widgets()
        self._setup_layout()

    def _setup_styles(self):
        style = ttk.Style()
        style.configure('TFrame', background='#f0f0f0')
        style.configure('Header.TLabel',
                        font=('Helvetica', 16, 'bold'),
                        padding=10,
                        background='#f0f0f0')
        style.configure('Status.TLabel',
                        font=('Helvetica', 10),
                        padding=5,
                        background='#f0f0f0')
        style.configure('Action.TButton',
                        font=('Helvetica', 10, 'bold'),
                        padding=5)
        style.configure('TCheckbutton',
                        background='#f0f0f0',
                        font=('Helvetica', 10))
        style.configure('FileType.TLabelframe.Label',
                        font=('Helvetica', 12, 'bold'))

    def _create_widgets(self):
        self.main_container = ttk.Frame(self)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.header = ttk.Label(
            self.main_container,
            text="Document Processing System",
            style='Header.TLabel'
        )

        self.dir_frame = ttk.Frame(self.main_container)
        self.dir_label = ttk.Label(
            self.dir_frame,
            text="Select Directory:",
            style='Status.TLabel'
        )
        self.dir_entry = ttk.Entry(self.dir_frame, width=50)
        self.browse_button = ttk.Button(
            self.dir_frame,
            text="Browse",
            style='Action.TButton',
            command=self._browse_directory
        )

        self.file_type_frame = ttk.LabelFrame(self.main_container, text="File Types to Process", style='FileType.TLabelframe')

        self.checkbuttons = {}
        for file_type, var in self.file_types.items():
            cb = ttk.Checkbutton(self.file_type_frame, text=file_type, variable=var, style='TCheckbutton')
            cb.pack(side=tk.LEFT, padx=10, pady=5)
            self.checkbuttons[file_type] = cb

        self.status_frame = ttk.Frame(self.main_container)
        self.status_label = ttk.Label(
            self.status_frame,
            text="Ready",
            style='Status.TLabel'
        )

        self.progress_frame = ttk.Frame(self.main_container)
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='determinate',
            length=300
        )
        self.progress_label = ttk.Label(
            self.progress_frame,
            text="0%",
            style='Status.TLabel'
        )

        self.monitor_frame = ttk.Frame(self.main_container)
        self.cpu_label = ttk.Label(
            self.monitor_frame,
            text="CPU: 0%",
            style='Status.TLabel'
        )
        self.memory_label = ttk.Label(
            self.monitor_frame,
            text="Memory: 0%",
            style='Status.TLabel'
        )

        self.log_frame = ttk.Frame(self.main_container)
        self.log_text = scrolledtext.ScrolledText(
            self.log_frame,
            height=10,
            width=70
        )

        self.button_frame = ttk.Frame(self.main_container)
        self.process_button = ttk.Button(
            self.button_frame,
            text="Start Processing",
            style='Action.TButton',
            command=self._start_processing
        )
        self.cancel_button = ttk.Button(
            self.button_frame,
            text="Cancel",
            style='Action.TButton',
            state=tk.DISABLED
        )

        self.log_handler = LogHandler(self.log_text)
        logging.getLogger().addHandler(self.log_handler)
        logging.getLogger().setLevel(logging.INFO)
        self._start_monitoring()

    def _setup_layout(self):
        self.main_container.columnconfigure(0, weight=1)
        self.header.pack(fill=tk.X, pady=(0, 10))

        self.dir_frame.pack(fill=tk.X, pady=5)
        self.dir_label.pack(side=tk.LEFT, padx=5)
        self.dir_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.browse_button.pack(side=tk.LEFT, padx=5)

        self.file_type_frame.pack(fill=tk.X, pady=10)

        self.status_frame.pack(fill=tk.X, pady=5)
        self.status_label.pack(side=tk.LEFT, padx=5)

        self.progress_frame.pack(fill=tk.X, pady=5)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.progress_label.pack(side=tk.LEFT, padx=5)

        self.monitor_frame.pack(fill=tk.X, pady=5)
        self.cpu_label.pack(side=tk.LEFT, padx=5)
        self.memory_label.pack(side=tk.LEFT, padx=5)

        self.log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5)

        self.button_frame.pack(fill=tk.X, pady=5)
        self.process_button.pack(side=tk.LEFT, padx=5)
        self.cancel_button.pack(side=tk.LEFT, padx=5)

    def _browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.dir_entry.delete(0, tk.END)
            self.dir_entry.insert(0, directory)

    def _start_processing(self):
        directory = self.dir_entry.get()
        if not directory:
            messagebox.showerror("Error", "Please select a directory")
            return

        self.process_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)
        self.status_label.config(text="Processing...")

        Thread(target=lambda: asyncio.run(self._process_documents(directory))).start()

    def _cancel_processing(self):
        self.status_label.config(text="Cancelling...")
        # Implementation for cancellation logic

    def _start_monitoring(self):
        def update_monitoring():
            cpu_percent = psutil.cpu_percent()
            memory_percent = psutil.virtual_memory().percent

            self.cpu_label.config(text=f"CPU: {cpu_percent}%")
            self.memory_label.config(text=f"Memory: {memory_percent}%")

            self.after(1000, update_monitoring)

        update_monitoring()

    def update_progress(self, current, total):
        percentage = int((current / total) * 100)
        self.progress_bar["value"] = percentage
        self.progress_label.config(text=f"{percentage}%")
        self.update_idletasks()

    async def _process_documents(self, directory):
        try:
            processor = DocumentProcessor(self.processing_config, self.azure_config)
            selected_file_types = [DocumentType[file_type].value for file_type, var in self.file_types.items() if var.get()]
            results = await processor.process_directory(Path(directory), self.update_progress, selected_file_types, self.update_status, batch_size=10)

            self.status_label.config(text="Completed! Results saved to output directory.")
            self.process_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            logging.info("Processing completed. Results exported to output directory.")

            messagebox.showinfo("Success", "Processing completed! Results saved to output directory.")

        except PermissionError as e:
            self.status_label.config(text="Permission error")
            self.process_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            logging.error(f"Permission error: {e}")
            messagebox.showerror("Permission Error", f"Permission error: {str(e)}")
        except Exception as e:
            self.status_label.config(text="Error occurred")
            self.process_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            logging.error(f"Processing error: {e}")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

    def update_status(self, file_name):
        self.status_label.config(text=f"Processing file: {file_name}")
        self.update_idletasks()

class ResultsExporter:
    @staticmethod
    async def to_excel(batch_results: List[Dict[str, Any]], output_path: Path, analyzer: ContentAnalyzer, update_progress):
        try:
            flattened_results = []
            total_tasks = len(batch_results)
            progress = 0

            def progress_callback():
                nonlocal progress
                progress += 1
                update_progress(progress, total_tasks)

            formatted_infos = []
            for result in batch_results:
                for analysis in result.get('analyses', []):
                    logging.debug(f"Processing analysis: {analysis}")
                    analysis_type = analysis.get('analysis_type')
                    result_content = analysis.get('result')
                    if analysis_type is None:
                        logging.error(f"Missing analysis_type in analysis: {analysis}")
                        continue
                    formatted_info = analyzer.format_information(result_content, analysis_type)
                    formatted_infos.append((formatted_info, result, analysis))

            for formatted_info, result, analysis in formatted_infos:
                analysis_type = analysis.get('analysis_type')
                logging.debug(f"Flattening result for analysis_type: {analysis_type}")
                result_content = analysis.get('result')
                flattened_result = {
                    'file_path': result.get('file_path'),
                    'metadata': str(result.get('metadata')),  # Convert metadata to string
                    'analysis_type': analysis_type,
                    'result': str(result_content),  # Ensure result content is a string
                    'start_date': formatted_info.get('start_date'),
                    'end_date': formatted_info.get('end_date'),
                    'contract_duration': formatted_info.get('contract_duration'),
                    'key_deliverables': str(formatted_info.get('key_deliverables')),  # Convert list to string
                    'vendor_capabilities': str(formatted_info.get('vendor_capabilities')),
                    'risk_factors': str(formatted_info.get('risk_factors')),
                    'compliance_status': str(formatted_info.get('compliance_status')),
                    'past_performance_metrics': str(formatted_info.get('past_performance_metrics')),
                    'key_strengths_weaknesses': str(formatted_info.get('key_strengths_weaknesses')),
                    'technical_requirements': str(formatted_info.get('technical_requirements')),
                    'performance_metrics': str(formatted_info.get('performance_metrics')),
                    'compatibility_requirements': str(formatted_info.get('compatibility_requirements')),
                    'implementation_timeline': str(formatted_info.get('implementation_timeline')),
                    'maintenance_requirements': str(formatted_info.get('maintenance_requirements'))
                }
                flattened_results.append(flattened_result)
                progress_callback()

            new_data = pd.DataFrame(flattened_results)
            logging.debug(f"Flattened results: {flattened_results}")
            logging.debug(f"New data DataFrame: {new_data}")

            if output_path.exists():
                wb = load_workbook(output_path)
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                    for sheet_name, data in new_data.groupby('analysis_type'):
                        logging.debug(f"Appending data to sheet: {sheet_name}")
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            for r in dataframe_to_rows(data, index=False, header=False):
                                ws.append(r)
                        else:
                            data.to_excel(writer, sheet_name=sheet_name, index=False)
                wb.save(output_path)
            else:
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                    for sheet_name, data in new_data.groupby('analysis_type'):
                        data.to_excel(writer, sheet_name=sheet_name, index=False)

            logging.info(f"Results exported to {output_path}")
        except Exception as e:
            logging.error(f"Error in to_excel: {e}")
            raise

def main():
    # Enhanced Logging Setup
    logging.basicConfig(level=logging.DEBUG)
    logger = logging.getLogger()
    text_widget = None  # Replace with your actual text widget
    if text_widget:
        handler = LogHandler(text_widget)
        logger.addHandler(handler)
   
    # Ensure output directory exists
    output_dir = ProcessingConfig.OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)

    # Start the application
    app = EnhancedDocumentProcessorApp()
   
    # Configure window styling
    app.option_add('*TButton*Padding', 5)
    app.option_add('*TButton*Font', 'Helvetica 10')
    app.option_add('*TLabel*Font', 'Helvetica 10')
   
    # Center the window on screen
    window_width = 800
    window_height = 600
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()
    center_x = int(screen_width/2 - window_width/2)
    center_y = int(screen_height/2 - window_height/2)
    app.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    app.mainloop()

if __name__ == "__main__":
    main()
